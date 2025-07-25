from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException,NoAlertPresentException,UnexpectedAlertPresentException,NoSuchElementException,ElementClickInterceptedException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import random



#from selenium.webdriver.firefox.service import Service
#from selenium.webdriver.firefox.options import Options
#from selenium.webdriver.common.by import By
#from selenium.webdriver.common.action_chains import ActionChains
#from selenium.webdriver.support.ui import WebDriverWait

import time
import openpyxl
import random

import os
print(os.getcwd())

file1="C:/Users/hp/Desktop/khasra_1430.xlsx"
file2="C:/Users/hp/Desktop/data.xlsx"
workbook1=openpyxl.load_workbook(file1)
workbook2=openpyxl.load_workbook(file2)
sheet1=workbook1["chatarpur_1430"] # select village from excel sheet
sheet2=workbook2["credentials"]

row_count=sheet1.max_row+1
col_count=sheet1.max_column
# print(row_count)
# print(col_count)

gata=[sheet1.cell(i,1).value for i in range (3,row_count)]

raw_gata_id = [sheet1.cell(i,2).value for i in range (3,row_count)] # gata id contains xao cleaning required

gata_id = list(map(lambda x:x.strip('\xa0'),raw_gata_id))

# fasal=[sheet1.cell(i,11).value for i in range (3,row_count)]
fasal = list(map(lambda i: sheet1.cell(i, 16).value, range(3, row_count)))
print(fasal)



fasal=[]
boring_type=[]
iterable_area=[]
agri_nonagri =[]
vivran = []
gata_area = []
temp_gata_list = []
gata_name = []
for i in range(3,row_count):
    fasal.append(sheet1.cell(i,16).value)
    boring_type.append(sheet1.cell(i,10).value)
    gata_area.append(sheet1.cell(i,3).value)
    agri_nonagri.append(sheet1.cell(i,21).value)
    vivran.append(sheet1.cell(i,22).value)
    gata_name.append(sheet1.cell(i,5).value)
print(gata_name)
print(gata_area)



# print(fasal)
# print(boring_type)
# print(agri_nonagri)
# print(vivran)

#------VARIABLES
GATA_TYPE=""
FASAL ={}
SICHAI_VIDHI=""



halka_name=sheet2.cell(2,1).value
pass_word=sheet2.cell(2,2).value
gram_name=sheet2.cell(2,3).value
total_gata=sheet2.cell(2,4).value
start_gata=sheet2.cell(2,5).value




serv_obj=Service("chromedriver.exe")
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver=webdriver.Chrome(service=serv_obj)
actions=ActionChains(driver)
mywait=WebDriverWait(driver,10)
actions = ActionChains(driver)

# Set up Firefox driver
#serv_obj = Service("geckodriver.exe")  # geckodriver is for Firefox
#firefox_options = Options()
#firefox_options.set_preference("detach", True)  # This keeps browser open after script ends (not always needed)
#driver = webdriver.Firefox(service=serv_obj, options=firefox_options)
#Initialize ActionChains and WebDriverWait
#actions = ActionChains(driver)
#mywait = WebDriverWait(driver, 10)



row_count=sheet1.max_row
col_count=sheet1.max_column

zoom_level = 65
script = f"document.body.style.zoom='{zoom_level}%'"
driver.execute_script(script)


def load_first_page():
    driver.get("https://ekhasra.up.gov.in/khasra_n/#/SelectLogin")
    time.sleep(1)
    driver.find_element(By.XPATH,"/html/body/app-root/selectlogin/div[2]/div/div[6]/div/div[4]/a").click()
    time.sleep(1)

def login_page():
    selectDistrict = Select(driver.find_element(By.ID, "up_district"))
    selectDistrict.select_by_visible_text("रामपुर")
    time.sleep(1)
    selectTehsil = Select(driver.find_element(By.ID, "up_tehsil"))
    selectTehsil.select_by_visible_text("स्वार")
    
    
    time.sleep(1.5)
    selecthalka = Select(driver.find_element(By.ID, "up_halka"))
    time.sleep(3)
    selecthalka.select_by_index(44)
    #captcha_value = driver.find_element(By.ID, "CaptchaDiv").text
    #driver.find_element(By.ID, "CaptchaInput").send_keys(captcha_value)
    print(driver.find_element(By.ID,"captcha").text)
    driver.find_element(By.ID, "password").send_keys(pass_word)
    input("enter captcha")
    driver.find_element(By.XPATH,"/html/body/app-root/login/div/div/div/form/div[6]/button").click()
    time.sleep(2)

def load_third_page():
    Select(driver.find_element(By.XPATH,"//*[@id=\"fasalYear\"]")).select_by_visible_text("1432 (1 जुलाई 2024 से 30 जून 2025)")
    time.sleep(1)
    Select(driver.find_element(By.XPATH,"//*[@id=\"myFasal\"]")).select_by_index(3)
    time.sleep(1)
    Select(driver.find_element(By.XPATH,"//*[@id=\"gram_name\"]")).select_by_index(5) # enter index of village name 
    time.sleep(3)
    driver.find_element(By.XPATH,"/html/body/app-root/lekhpalhome/div[3]/div[2]/div/div/form/div[4]/button").click()
    

def fourth_page():
    time.sleep(3)
    driver.find_element(By.XPATH,"/html/body/app-root/lekhpaldashboard/div/div[2]/div/div/div/div[3]/div[3]/button").click()
    

def gata_display():
    time.sleep(1)
    driver.find_element(By.XPATH,"/html/body/app-root/lekhpalgatalist/div[1]/div[3]/table/tbody/tr[1]/td[5]/button").click()
           
    
    
       

            

def feeding():
    akrishak_selection_index = ""
    try:
     driver.find_element(By.XPATH,"/html/body/app-root/lekhpalgatalist/div[2]/div/div/div[2]/div/div[5]/div[2]/button").click()
     
    except:
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalgatalist/div[2]/div/div/div[2]/div/div[4]/div[2]/button").click()
    #-----gata vivran
    time.sleep(.5)
    time.sleep(.5)
    driver.find_element(By.XPATH,"/html/body/app-root/lekhpalownerdetails/div/div[3]/div/div[2]/button").click()
    time.sleep(1)
    
    driver.find_element(By.XPATH,"//*[@id=\"option1\"]").click()
    # gata_properties
    temp_gata_id = driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[3]/table/tbody/tr/td[3]").text
    temp_gata_area = driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[3]/table/tbody/tr/td[4]").text
    temp_gata = driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[3]/table/tbody/tr/td[2]").text
    
    print(temp_gata,temp_gata_id,temp_gata_area)
    time.sleep(.5)

    
    #matching
    
    attribute_index = gata_id.index(temp_gata_id)
    print(attribute_index)
    print(agri_nonagri[attribute_index])
    print(fasal[attribute_index])




    print(boring_type[attribute_index])
    print("#-----")
    print(gata[attribute_index])
    print(temp_gata_area)
    print(gata_area[attribute_index])
    
    driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[5]/div[2]/button").click()
    #fasal vivran
    if agri_nonagri[attribute_index] == "कृषित" and gata_area[attribute_index] == float(temp_gata_area):
        # akrishak vivran
        time.sleep(1.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[4]/div/button").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/div/button").click()
        time.sleep(1)
        
        #fallow land
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/form/div[3]/div/button").click()
        time.sleep(3)
        
        # tree details
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/form/div[1]/div/button").click()
        time.sleep(1.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[1]/form/div[4]/div/button[2]").click()
        time.sleep(.5)
        
        # main crop
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[2]/div/div/div[2]/form/div/select/option[3]").click()
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[2]/div/div/div[3]/button").click()
        time.sleep(.5)
        
        # boarder crop
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[1]/form/div[4]/div/button[2]").click()
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[2]/div/div/div[2]/form/div/select/option[2]").click()
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[2]/div/div/div[3]/button").click()
        time.sleep(3)
        
        # fasal details entry
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinserttree/div[1]/form/div[4]/div/button[2]").click()
        time.sleep(2)
        Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfasal/div[1]/div[6]/div[3]/select")).select_by_visible_text(fasal[attribute_index])
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfasal/div[1]/div[7]/div[2]/input").send_keys(gata_area[attribute_index])
        driver.find_element(By.XPATH,"//*[@id=\"flexSwitchCheckDefault\"]").click()
        time.sleep(.3)
        Select(driver.find_element(By.XPATH,"//*[@id=\"agriTech\"]")).select_by_visible_text(boring_type[attribute_index])
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfasal/div[1]/div[9]/div[2]/button").click()
        time.sleep(2)
        
        # tippani
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfasal/div[1]/div[9]/div[2]/button[2]").click()
        time.sleep(.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertremark/div/div[4]/div/div/textarea").send_keys("ok")
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertremark/div/div[5]/div[1]/button").click()
        
        # lock
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertremark/div/div[5]/div/button").click()
        
        # alert box
        time.sleep(2)
        driver.find_element(By.XPATH,"/html/body/div[3]/div/div[6]/button[1]").click()
        #driver.find_element(By.LINK_TEXT,"Close").click()
        
        
        
        
        

    if agri_nonagri[attribute_index] == "अकृषक/Non Agricultural"  and gata_area[attribute_index] == float(temp_gata_area):
        time.sleep(1.5)
        
        if gata_name[attribute_index] == " रास्ता" or gata_name[attribute_index] == " चकमार्ग" or gata_name[attribute_index] == " चकरोड":
            akrishak_selection_index = "28"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)
        elif gata_name[attribute_index] == " गूल" or gata_name[attribute_index] == gata_name[attribute_index] == " तालाब" or gata_name[attribute_index] == " नदी" or gata_name[attribute_index] ==" झील":
            akrishak_selection_index = "17"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        elif gata_name[attribute_index] == " आबादी":
            akrishak_selection_index = "3"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        elif gata_name[attribute_index] == " खलियान":
            akrishak_selection_index = "12"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        elif gata_name[attribute_index] == " खाद के गड्डे":
            akrishak_selection_index = "30"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        elif gata_name[attribute_index] == " देवस्थान":
            akrishak_selection_index = "32"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        elif gata_name[attribute_index] == " पुरानी परती" or gata_name[attribute_index] == " बंजर" or gata_name[attribute_index]== " नवीन परती":
            akrishak_selection_index = "46"
            Select(driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/select")).select_by_value(akrishak_selection_index)

        else:
            input("select manually")


        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[3]/input").send_keys(gata_area[attribute_index])
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[4]/div/button").click()
        time.sleep(.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[1]/div/div/button").click()
        time.sleep(.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalownerdetails/div/div[3]/div/div[2]/button[2]").click()
        time.sleep(2)
        driver.find_element(By.XPATH,"/html/body/div[3]/div/div[6]/button[1]").click()
   


    if agri_nonagri[attribute_index] == "आकृषित/Fallow Land"  and gata_area[attribute_index] == float(temp_gata_area):
        #input("feed manually as fallow land")
        time.sleep(1.5)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[4]/div/button").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/form/div[1]/div/button").click()
        time.sleep(2)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/form/div[2]/input").clear()
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/form/div[2]/input").send_keys(gata_area[attribute_index])
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/form/div[3]/div/button").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalinsertfallow/div/div[1]/div/div/button").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalcroppingpattern/div/div[1]/div[3]/button").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/lekhpalownerdetails/div/div[3]/div/div[2]/button[2]").click()
        time.sleep(2)
        driver.find_element(By.XPATH,"/html/body/div[3]/div/div[6]/button[1]").click()
        
        
        




load_first_page()
login_page()
load_third_page()
fourth_page()

while True:
    try:
        gata_display()
        feeding()
    except Exception as e:
        print(f"error occured:{e}")
        input("please correct error")


